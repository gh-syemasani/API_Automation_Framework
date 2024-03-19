import openpyxl
import requests
import json
from requests.auth import HTTPBasicAuth
import os
from tests.conftest import testcase_logger
import pytest
import time
from urllib.parse import parse_qs
import base64
from secrets_manager import SecretsManager

secret_manager_client = SecretsManager()

# atlassian_url = secret_manager_client.get_secret('autodoc/atlassian/url')
apiusername = secret_manager_client.get_secret('/sqa/security/mulesoft_env/apitesting_username')
apipassword = secret_manager_client.get_secret('/sqa/security/mulesoft_env/apitesting_password')

apiclientid = secret_manager_client.get_secret('/sqa/security/mulesoft_env/apitesting_client_id	')
apiclientsecret = secret_manager_client.get_secret('/sqa/security/mulesoft_env/apitesting_client_secret')

# import logging
# logging.basicConfig(level=logging.DEBUG)
# logger = logging.getLogger(__name__)



def write_test_results(file_path, sheet_name, test_data_list):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row_index, test_data in enumerate(test_data_list, start=2):
        
        for sheet_row_index, sheet_test_case in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
            if sheet_test_case[0] == test_data["test_case"]:
                row_index = sheet_row_index
                break

        sheet.cell(row=row_index, column=sheet.max_column - 1, value=test_data["actual_response"])
        sheet.cell(row=row_index, column=sheet.max_column, value=test_data["result"])

    # logger.info(f"Processing test case {test_data['test_case']}")
    workbook.save(file_path)
    time.sleep(5)

        

def read_test_data(file_path, sheet_name):
    test_data = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        (
            test_case, execute, http_method, auth_type, authorization_header_json,
            endpoint, request_type, request_data_type, request_data, expected_status_code,
            expected_response,excluded_fields_cell,actual_response,result
        ) = row

        if execute.lower() == "yes":
            authorization_header = json.loads(authorization_header_json)

            if request_data:
                try:
                    request_data = json.loads(request_data)
                except json.JSONDecodeError:
                    request_data = None

            # Split excluded fields by comma
            excluded_fields = [field.strip() for field in excluded_fields_cell.split(',')] if excluded_fields_cell else []

            test_data.append({
                "test_case": test_case,
                "http_method": http_method,
                "auth_type": auth_type,
                "authorization_header": authorization_header,
                "endpoint": endpoint,
                "request_type": request_type,
                "request_data_type": request_data_type,
                "request_data": request_data,
                "expected_status_code": int(expected_status_code),
                "expected_response": json.loads(expected_response),
                "excluded_fields": excluded_fields,  # List of excluded fields
                "actual_response": actual_response,
                "result": result
                
            })
        print(test_data)
    return test_data


API_BASE_URL = "https://mule-nonprod-bizapps.guardanthealth.com/"
SERVICE_BASE_URL = "Service URL"

def perform_get_request(url, headers):
    response = requests.get(url, headers=headers)
    return response

def perform_post_request(url, headers, request_data, request_data_type,auth_data):
    if request_data_type == "JSON":
        print(json.dumps(request_data))
        print(headers)
        # logger.debug(f"Request Headers: {headers}")
        # logger.debug(f"Request Payload: {json.dumps(request_data)}")
        response = requests.post(url, json=request_data, auth=(auth_data["client_id"], auth_data["client_secret"]), headers=headers)
        # response = requests.post(url, json=request_data, headers=headers)
    elif request_data_type == "Form Data":
        data = parse_qs(request_data)
        print(headers)
        print(data)
        # logger.debug(f"Request Headers: {headers}")
        # logger.debug(f"Request Payload: {data}")
        response = requests.post(url, data=data, headers=headers)
    elif request_data_type == "XML":
        headers['Content-Type'] = 'application/xml'
        print(headers)
        print(request_data)
        # logger.debug(f"Request Headers: {headers}")
        # logger.debug(f"Request Payload: {request_data}")
     
        response = requests.post(url, data=request_data, headers=headers)
    else:
        raise ValueError(f"Unsupported request Data Type: {request_data_type}")
    return response

def perform_put_request(url, headers, request_data, request_data_type):
    if request_data_type == "JSON":
        response = requests.put(url, json=request_data, headers=headers)
    elif request_data_type == "Form Data":
        data = parse_qs(request_data)
        response = requests.put(url, data=data, headers=headers)
    elif request_data_type == "XML":
        headers['Content-Type'] = 'application/xml'
        response = requests.put(url, data=request_data, headers=headers)
    else:
        raise ValueError(f"Unsupported request data type: {request_data_type}")
    return response

def perform_delete_request(url, headers):
    response = requests.delete(url, headers=headers)
    return response

def validate_response(response, expected_status_code, expected_response, excluded_fields=None):
    status_code_match = response.status_code == expected_status_code

    if excluded_fields is None:
        excluded_fields = []

    if status_code_match:
        actual_response = response.json()

        # Exclude fields and their keys with values from both actual and expected responses
        for field in excluded_fields:
            actual_data = actual_response.get("data", {})
            expected_data = expected_response.get("data", {})

            # Remove the field and its key with value
            if field in actual_data:
                actual_data.pop(field, None)
            if field in expected_data:
                expected_data.pop(field, None)

        # Compare the modified "data" part of the expected and actual responses
        expected_data = expected_response.get("data", {})
        actual_data = actual_response.get("data", {})

        print("expected_data") 
        print(expected_data) 

        print("actual_data")
        print(actual_data)


        response_match = expected_data == actual_data
        return response_match

    return False


# def validate_response(response, expected_status_code, expected_response):
#     status_code_match = response.status_code == expected_status_code
#     # response_match = response.json() == expected_response
#     # return status_code_match and response_match
#     return status_code_match

def get_authorization_headers(auth_type, auth_data):
    headers = {"Content-Type": "application/json"}

    if auth_type == "Bearer":
        headers["Authorization"] = f"Bearer {auth_data['token']}"
    elif auth_type == "Custom1":
          if "username" in auth_data and "password" in auth_data:
           usern = auth_data["username"]
           password = auth_data["password"]

# Create an instance of HTTPBasicAuth
           credentials = base64.b64encode(f"{usern}:{password}".encode("utf-8")).decode("utf-8")

# Update the headers with the correct 'Authorization', 'client_id', and 'client_secret'
           headers.update({
                "Authorization": f"Basic {credentials}",
                "client_id":apiclientid,
                "client_secret":apiclientsecret
                # "client_id": auth_data.get("client_id", ""),
                # "client_secret": auth_data.get("client_secret", "")
})

    elif auth_type == "Basic":
        if "username" in auth_data and "password" in auth_data:
            auth_header = HTTPBasicAuth(auth_data["username"], auth_data["password"])
        elif "client_id" in auth_data and "client_secret" in auth_data:
            #   client_id=auth_data["client_id"]
            #   client_secret=auth_data["client_secret"]
            #   credentials = f"{client_id}:{client_secret}"
            #   base64_credentials = base64.b64encode(credentials.encode("utf-8")).decode("utf-8")
            #   auth_header=  {
            #        "Content-Type": "application/json",
            #        "Authorization": "Basic" base64_credentials
            #     } 

            # auth = HTTPBasicAuth(auth_data["client_id"], auth_data["client_secret"])
            # auth_header = {"Content-Type": "application/json", "Authorization": auth}
            # auth_header = HTTPBasicAuth(auth_data["client_id"],auth_data["client_secret"])
            # client_id=apiclientid
            # client_secret=apiclientsecret
            
            auth_header = HTTPBasicAuth(auth_data[apiclientid],auth_data[apiclientsecret])
            
            print(auth_header)
        elif "username" in auth_data or "password" in auth_data:
            raise ValueError("Incomplete credentials for Basic authentication. Both username and password are required.")
        else:
            raise ValueError("Missing credentials for Basic authentication. Either username and password or client_id and client_secret are required.")

        headers["Authorization"] = f"Basic {auth_header.username}:{auth_header.password}"
    elif auth_type == "Custom":
        headers.update({
              "client_id":apiclientid,
              "client_secret":apiclientsecret

            # "client_id": auth_data.get("client_id", ""),
            # "client_secret": auth_data.get("client_secret", "")
        })
    elif auth_type == "API Key":
        headers["Authorization"] = f"ApiKey {auth_data.get('api_key', '')}"
    elif auth_type == "OAuth1.0":
        pass
    elif auth_type == "OAuth2.0":
        pass
    elif auth_type == "ClientCredentials" or auth_type =="Custom1":
            
            
            headers.update({

             "client_id":apiclientid ,
             "client_secret": apiclientsecret
        
            #  "client_id": auth_data.get("client_id", ""),
            # "client_secret": auth_data.get("client_secret", "")
            })
    elif auth_type == "UsernamePassword":
     
        headers.update({

              "client_id":apiclientid ,
              "client_secret": apiclientsecret,
               "username":apiusername,
               "password":apipassword

        #     "client_id": auth_data.get("client_id", ""),
        #     "client_secret": auth_data.get("client_secret", ""),
        #     "username": auth_data.get("username", ""),
        #     "password": auth_data.get("password", "")
         })
    else:
        raise ValueError(f"Unsupported Authorization Type: {auth_type}")

    return headers

Excel_Requestfile = os.path.join(os.path.dirname(__file__), '.', 'test_data', 'API_Test_Data.xlsx')

@pytest.mark.parametrize("test_data", read_test_data(Excel_Requestfile, "Input_Result"))
def test_api_requests(test_data, write_to_excel=True):
    if test_data["request_type"] == "API":
        base_url = API_BASE_URL
    elif test_data["request_type"] == "Service":
        base_url = SERVICE_BASE_URL
    else:
        raise ValueError(f"Unsupported Request Type: {test_data['request_type']}")

    url = base_url + test_data["endpoint"]
    http_method = test_data["http_method"].upper()
    request_data = test_data["request_data"]
    request_data_type = test_data["request_data_type"]

    headers = get_authorization_headers(test_data["auth_type"], test_data.get("authorization_header", {}))

    print(f"Running test case: {test_data['test_case']}")
    print(f"URL: {url}")
    print(f"HTTP Method: {http_method}")
    print(f"Headers: {headers}")
    print(f"Request Data: {request_data}")

    if http_method == "GET":
        response = perform_get_request(url, headers)
    elif http_method == "POST":
        auth_data = {"client_id": "your_client_id", "client_secret": "your_client_secret"}
        response = perform_post_request(url, headers, request_data, request_data_type, auth_data)
    elif http_method == "PUT":
        response = perform_put_request(url, headers, request_data, request_data_type)
    elif http_method == "DELETE":
        response = perform_delete_request(url, headers)
    else:
        raise ValueError(f"Unsupported HTTP method: {http_method}")
    
    excluded_fields_cell = test_data.get("excluded_fields", None)

    # Call validate_response with excluded fields
    test_result = validate_response(response, test_data["expected_status_code"], test_data["expected_response"], excluded_fields_cell)


    # test_result = validate_response(response, test_data["expected_status_code"], test_data["expected_response"])

    print(f"Test result: {test_result}")

    # Update the test result and actual response in the test_data dictionary
    test_data["result"] = "Pass" if test_result else "Fail"
    test_data["actual_response"] = response.text

    # Always write the test results to the Excel file
    if write_to_excel:
        print(f"Writing test results to Excel for test case: {test_data['test_case']}")
        write_test_results(Excel_Requestfile, "Input_Result", [test_data])

    assert test_result, f"Test case {test_data['test_case']} failed"


if __name__ == "__main__":
    Excel_Requestfile = os.path.join(os.path.dirname(__file__), '.', 'test_data', 'API_Test_Data.xlsx')
    
    test_data_list = read_test_data(Excel_Requestfile, "Input_Result")

    for test_data in test_data_list:
        test_api_requests(test_data)
        # write_test_results(Excel_Requestfile, "Input_Result", [test_data])